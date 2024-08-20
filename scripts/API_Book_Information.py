import httpx
import asyncio
import openpyxl

# API 정보 
auth_key = '''YOUR API KEY'''
input_xlsx = '/kaggle/input/1850-ver/popular_books_with_analysis.xlsx'
output_xlsx = '/kaggle/working/popular_books_with_analysis.xlsx'

async def book_analysis(auth_key, isbn):
    base_url = 'http://data4library.kr/api/usageAnalysisList'
    url = f"{base_url}?authKey={auth_key}&isbn13={isbn}&format=json"
    print(f"\n[Request URL] {url}\n")
    
    async with httpx.AsyncClient(timeout=20.0) as client:
        try:
            response = await client.get(url)
            response.raise_for_status()
            data = response.json()
            
            analysis_data = {"단어": [], "가중치": [], "마니아 추천 도서": []}

            if 'response' in data and 'book' in data['response']:
                original_book_name = data['response']['book'].get('bookname', 'N/A')
                print(f"Original Book: {original_book_name}\n")
                
            if 'response' in data and 'keywords' in data['response']:
                # 단어 및 가중치 추출 (최대 10개)
                for word in data['response']['keywords'][:10]:
                    analysis_data["단어"].append(word['keyword'].get('word', 'N/A'))
                    analysis_data["가중치"].append(str(word['keyword'].get('weight', 'N/A')))
                    
            if 'response' in data and 'maniaRecBooks' in data['response']:
                # 마니아를 위한 추천도서 리스트 확인 (최대 3권)
                recommended_books = data['response']['maniaRecBooks'][:3]
                for book in recommended_books:
                    book_info = {
                        "도서명": book['book'].get('bookname', 'N/A'),
                        "저자명": book['book'].get('authors', 'N/A'),
                        "출판사": book['book'].get('publisher', 'N/A'),
                        "출판년도": book['book'].get('publication_year', 'N/A'),
                        "ISBN": book['book'].get('isbn13', 'N/A')
                    }
                    analysis_data["마니아 추천 도서"].append(book_info)
                    
            return analysis_data
        except httpx.TimeoutException:
            print("요청 시간이 초과되었습니다.")
            return None
        except httpx.RequestError as e:
            print(f"요청 실패: {e}")
            return None

async def main():
    # 기존 Excel 파일 불러오기
    workbook = openpyxl.load_workbook(input_xlsx)
    sheet = workbook.active
    
    # 헤더 추가 (이미 추가되어 있을 경우 생략 가능)
    if sheet['I1'].value is None:
        sheet['I1'] = "단어"
    if sheet['J1'].value is None:
        sheet['J1'] = "가중치"
    
    headers = [
        "마니아 도서명1", "마니아 저자명1", "마니아 출판사1", "마니아 출판년도1", "마니아 ISBN1",
        "마니아 도서명2", "마니아 저자명2", "마니아 출판사2", "마니아 출판년도2", "마니아 ISBN2",
        "마니아 도서명3", "마니아 저자명3", "마니아 출판사3", "마니아 출판년도3", "마니아 ISBN3"
    ]
    
    for col_num, header in enumerate(headers, start=11):  # 열 'K'부터 시작
        if sheet.cell(row=1, column=col_num).value is None:
            sheet.cell(row=1, column=col_num).value = header
    
    # 5행부터 시작하여 8행까지 처리 예시 - sheet.iter_rows(min_row=5, max_row=8, min_col=6, max_col=6), start=5)
    for i, row in enumerate(sheet.iter_rows(min_row=1852, max_row=2001, min_col=6, max_col=6), start=1852):
        isbn_cell = row[0]
        if isbn_cell.value:
            analysis_data = await book_analysis(auth_key, isbn_cell.value)
            if analysis_data:
                # 단어와 가중치를 쉼표로 구분하여 엑셀에 저장
                sheet[f'I{isbn_cell.row}'] = ", ".join(analysis_data["단어"])
                sheet[f'J{isbn_cell.row}'] = ", ".join(analysis_data["가중치"])
                
                # 마니아 추천 도서 정보를 개별 열에 저장
                for idx, book in enumerate(analysis_data["마니아 추천 도서"], start=1):
                    sheet.cell(row=isbn_cell.row, column=10 + (idx - 1) * 5 + 1, value=book['도서명'])
                    sheet.cell(row=isbn_cell.row, column=10 + (idx - 1) * 5 + 2, value=book['저자명'])
                    sheet.cell(row=isbn_cell.row, column=10 + (idx - 1) * 5 + 3, value=book['출판사'])
                    sheet.cell(row=isbn_cell.row, column=10 + (idx - 1) * 5 + 4, value=book['출판년도'])
                    sheet.cell(row=isbn_cell.row, column=10 + (idx - 1) * 5 + 5, value=book['ISBN'])
    
    # 새로운 Excel 파일로 저장
    workbook.save(output_xlsx)
    print(f"Excel 파일로 저장 완료: {output_xlsx}")

# 현재 이벤트 루프가 이미 실행 중인지 확인하고, 아니면 새로 실행
loop = asyncio.get_event_loop()
if loop.is_running():
    # 현재 실행 중인 루프에서 직접 실행
    task = loop.create_task(main())
    await task
else:
    # 새로운 루프를 실행
    asyncio.run(main())
